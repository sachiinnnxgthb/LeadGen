"""Excel CRM exporter.

Builds the styled multi-sheet CRM workbook (and single-segment workbooks) with
freeze panes, auto-filter, auto-fit columns, hyperlinks, conditional formatting,
and a charted Dashboard sheet.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from lead_intel.core.logging import get_logger
from lead_intel.domain.enums import WebsiteQuality, WebsiteStatus
from lead_intel.domain.models import Lead
from lead_intel.exporters import excel_styles as styles
from lead_intel.exporters.analytics import DashboardStats
from lead_intel.exporters.rows import (
    COLUMNS,
    HYPERLINK_COLUMNS,
    WIDE_TEXT_COLUMNS,
    LeadRow,
    lead_to_row,
)

logger = get_logger("exporters.excel")

_MAX_WIDTH = 40
_WIDE_WIDTH = 55
_SCORE_COL = "Website Score"
_LEAD_COL = "Lead Score"
_PRIORITY_COL = "Priority"


def _col_index(name: str) -> int:
    """1-based column index of a named column."""
    return COLUMNS.index(name) + 1


def write_leads_sheet(ws: Worksheet, leads: list[Lead], *, title: str) -> None:
    """Render a fully-styled lead table onto ``ws``."""
    ws.title = title
    rows = [lead_to_row(lead) for lead in leads]

    # Header.
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = styles.HEADER_FONT
        cell.fill = styles.HEADER_FILL
        cell.alignment = styles.HEADER_ALIGN

    # Data.
    for r, row in enumerate(rows, start=2):
        shade = styles.zebra_fill(r)
        for col_idx, name in enumerate(COLUMNS, start=1):
            value = row.values[name]
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.alignment = styles.WRAP_ALIGN if name in WIDE_TEXT_COLUMNS else styles.TOP_ALIGN
            cell.border = styles.CELL_BORDER
            if shade is not None:
                cell.fill = shade
            if name in HYPERLINK_COLUMNS and value:
                cell.hyperlink = str(value)
                cell.font = styles.LINK_FONT

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    _autofit_columns(ws, rows)
    if rows:
        _apply_conditional_formatting(ws, row_count=len(rows))


def _autofit_columns(ws: Worksheet, rows: list[LeadRow]) -> None:
    for col_idx, name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        longest = len(name)
        for row in rows:
            longest = max(longest, len(str(row.values[name])))
        cap = _WIDE_WIDTH if name in WIDE_TEXT_COLUMNS else _MAX_WIDTH
        ws.column_dimensions[letter].width = min(longest + 2, cap)


def _apply_conditional_formatting(ws: Worksheet, *, row_count: int) -> None:
    last = row_count + 1
    score_col = get_column_letter(_col_index(_SCORE_COL))
    lead_col = get_column_letter(_col_index(_LEAD_COL))
    prio_col = get_column_letter(_col_index(_PRIORITY_COL))

    ws.conditional_formatting.add(f"{score_col}2:{score_col}{last}", styles.website_score_rule())
    ws.conditional_formatting.add(f"{lead_col}2:{lead_col}{last}", styles.lead_score_rule())
    for rule in styles.priority_rules():
        ws.conditional_formatting.add(f"{prio_col}2:{prio_col}{last}", rule)


def write_dashboard_sheet(ws: Worksheet, stats: DashboardStats) -> None:
    """Render KPIs and charts onto the Dashboard sheet."""
    ws.title = "Dashboard"
    ws["A1"] = "Lead Intelligence Dashboard"
    ws["A1"].font = styles.TITLE_FONT
    ws.merge_cells("A1:D1")

    # KPI panel.
    row = 3
    for label, value in stats.as_kpi_pairs():
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = styles.KPI_LABEL_FONT
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = styles.KPI_VALUE_FONT
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    _write_distribution_charts(ws, stats, start_row=3)


def _write_distribution_charts(ws: Worksheet, stats: DashboardStats, *, start_row: int) -> None:
    """Write helper tables + a bar chart (categories) and pie chart (web status)."""
    # Category table (col D/E) feeding a bar chart.
    ws.cell(row=start_row - 1, column=4, value="Category").font = styles.KPI_LABEL_FONT
    ws.cell(row=start_row - 1, column=5, value="Count").font = styles.KPI_LABEL_FONT
    cat_items = list(stats.category_distribution.items())
    for i, (name, count) in enumerate(cat_items):
        ws.cell(row=start_row + i, column=4, value=name)
        ws.cell(row=start_row + i, column=5, value=count)

    if cat_items:
        bar = BarChart()
        bar.title = "Businesses by Category"
        bar.type = "col"
        bar.legend = None
        cat_end = start_row + len(cat_items) - 1
        data = Reference(ws, min_col=5, min_row=start_row - 1, max_row=cat_end)
        cats = Reference(ws, min_col=4, min_row=start_row, max_row=cat_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height, bar.width = 8, 16
        ws.add_chart(bar, "G3")

    # Website-status table (col D, lower) feeding a pie chart.
    status_row = start_row + len(cat_items) + 2
    status_data = [
        ("No Website", stats.without_website),
        ("Broken", stats.broken_websites),
        ("Outdated", stats.outdated_websites),
        ("Modern", stats.modern_websites),
    ]
    ws.cell(row=status_row - 1, column=4, value="Website Status").font = styles.KPI_LABEL_FONT
    ws.cell(row=status_row - 1, column=5, value="Count").font = styles.KPI_LABEL_FONT
    for i, (name, count) in enumerate(status_data):
        ws.cell(row=status_row + i, column=4, value=name)
        ws.cell(row=status_row + i, column=5, value=count)

    pie = PieChart()
    pie.title = "Website Status Mix"
    status_end = status_row + len(status_data) - 1
    labels = Reference(ws, min_col=4, min_row=status_row, max_row=status_end)
    data = Reference(ws, min_col=5, min_row=status_row - 1, max_row=status_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height, pie.width = 8, 12
    ws.add_chart(pie, "G20")


# -- workbook builders -----------------------------------------------------


def build_crm_workbook(leads: list[Lead], stats: DashboardStats) -> Workbook:
    """Build the full CRM workbook with all six sheets."""
    wb = Workbook()
    write_leads_sheet(wb.active, leads, title="All Leads")

    segments = [
        ("No Website", [x for x in leads if x.website_status == WebsiteStatus.NO_WEBSITE]),
        ("Outdated Website", [x for x in leads if x.website_quality == WebsiteQuality.OUTDATED]),
        ("Broken Website", [x for x in leads if x.website_quality == WebsiteQuality.BROKEN]),
        ("High Priority", [x for x in leads if x.is_high_priority]),
    ]
    for title, subset in segments:
        write_leads_sheet(wb.create_sheet(title), subset, title=title)

    write_dashboard_sheet(wb.create_sheet("Dashboard"), stats)
    return wb


def build_segment_workbook(leads: list[Lead], *, title: str) -> Workbook:
    """Build a single-sheet workbook for one lead segment."""
    wb = Workbook()
    write_leads_sheet(wb.active, leads, title=title)
    return wb


def build_dashboard_workbook(stats: DashboardStats) -> Workbook:
    """Build a standalone Dashboard workbook."""
    wb = Workbook()
    write_dashboard_sheet(wb.active, stats)
    return wb


def save_workbook(wb: Workbook, path: Path) -> Path:
    """Persist a workbook, creating parent directories as needed."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("wrote workbook", extra={"path": str(path)})
    return path
