"""Shared Excel styling.

Centralized fonts, fills, borders, and conditional-formatting rules so every
sheet in the workbook looks consistent and professional.
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Brand palette (hex, no leading '#').
NAVY = "1F3A5F"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F4F7"
GREEN = "C6EFCE"
GREEN_TEXT = "006100"
YELLOW = "FFEB9C"
YELLOW_TEXT = "9C6500"
RED = "FFC7CE"
RED_TEXT = "9C0006"
GREY = "D9D9D9"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=NAVY)
KPI_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True)

LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
TOP_ALIGN = Alignment(vertical="top")

_THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def zebra_fill(row_index: int) -> PatternFill | None:
    """Alternating row shading (``None`` for un-shaded rows)."""
    return PatternFill("solid", fgColor=LIGHT_GREY) if row_index % 2 == 0 else None


def website_score_rule() -> ColorScaleRule:
    """Red→Yellow→Green scale across 0-100."""
    return ColorScaleRule(
        start_type="num", start_value=0, start_color=RED,
        mid_type="num", mid_value=50, mid_color=YELLOW,
        end_type="num", end_value=100, end_color=GREEN,
    )


def lead_score_rule() -> ColorScaleRule:
    """Red→Yellow→Green scale across 0-10."""
    return ColorScaleRule(
        start_type="num", start_value=0, start_color=RED,
        mid_type="num", mid_value=5, mid_color=YELLOW,
        end_type="num", end_value=10, end_color=GREEN,
    )


def priority_rules() -> list[CellIsRule]:
    """Text-match fills for the Priority column."""
    return [
        CellIsRule(
            operator="equal", formula=['"High"'],
            fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TEXT, bold=True),
        ),
        CellIsRule(
            operator="equal", formula=['"Medium"'],
            fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_TEXT),
        ),
        CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=GREY)),
    ]
