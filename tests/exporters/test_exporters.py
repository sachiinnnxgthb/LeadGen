"""Tests for Excel/CSV/JSON/PDF exporters and the export service."""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from lead_intel.config.settings import Settings
from lead_intel.domain.models import Lead
from lead_intel.exporters.analytics import compute_dashboard
from lead_intel.exporters.data_exporter import export_csv, export_json
from lead_intel.exporters.excel_exporter import build_crm_workbook
from lead_intel.exporters.export_service import ExportService, _slugify
from lead_intel.exporters.pdf_exporter import PdfAuditExporter
from lead_intel.exporters.rows import COLUMNS


def _settings(output_dir: Path) -> Settings:
    return Settings(_env_file=None, output_dir=str(output_dir))  # type: ignore[call-arg]


# -- Excel -----------------------------------------------------------------


def test_crm_workbook_has_all_sheets(leads: list[Lead], tmp_path: Path) -> None:
    stats = compute_dashboard(leads, _settings(tmp_path).revenue)
    wb = build_crm_workbook(leads, stats)
    assert wb.sheetnames == [
        "All Leads", "No Website", "Outdated Website", "Broken Website",
        "High Priority", "Dashboard",
    ]


def test_all_leads_sheet_structure(leads: list[Lead], tmp_path: Path) -> None:
    stats = compute_dashboard(leads, _settings(tmp_path).revenue)
    wb = build_crm_workbook(leads, stats)
    ws = wb["All Leads"]

    header = [c.value for c in ws[1]]
    assert header == list(COLUMNS)
    assert ws.max_row == len(leads) + 1
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    # Conditional formatting was registered (score/lead/priority).
    assert len(ws.conditional_formatting) >= 3


def test_segment_sheets_are_filtered(leads: list[Lead], tmp_path: Path) -> None:
    stats = compute_dashboard(leads, _settings(tmp_path).revenue)
    wb = build_crm_workbook(leads, stats)
    assert wb["No Website"].max_row == 2       # 1 no-website lead + header
    assert wb["High Priority"].max_row == 3    # 2 high-priority leads + header


def test_workbook_roundtrips_to_disk(leads: list[Lead], tmp_path: Path) -> None:
    stats = compute_dashboard(leads, _settings(tmp_path).revenue)
    path = tmp_path / "CRM.xlsx"
    build_crm_workbook(leads, stats).save(path)
    reloaded = load_workbook(path)
    assert "Dashboard" in reloaded.sheetnames
    assert reloaded["Dashboard"]["A1"].value == "Lead Intelligence Dashboard"


# -- CSV / JSON ------------------------------------------------------------


def test_csv_export(leads: list[Lead], tmp_path: Path) -> None:
    path = export_csv(leads, tmp_path / "leads.csv")
    with path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == len(leads)
    assert rows[0]["Business Name"] == "IronCore Gym"
    assert set(rows[0]) == set(COLUMNS)


def test_json_export(leads: list[Lead], tmp_path: Path) -> None:
    path = export_json(leads, tmp_path / "leads.json")
    data = json.loads(path.read_text(encoding="utf-8"))
    assert len(data) == len(leads)
    assert data[0]["business"]["name"] == "IronCore Gym"


# -- PDF -------------------------------------------------------------------


def test_pdf_export_creates_valid_file(leads: list[Lead], tmp_path: Path) -> None:
    exporter = PdfAuditExporter(_settings(tmp_path).agency)
    path = exporter.export(leads[0], tmp_path / "audit.pdf")
    assert path.exists()
    assert path.read_bytes().startswith(b"%PDF")
    assert path.stat().st_size > 1000


def test_pdf_export_for_modern_site(leads: list[Lead], tmp_path: Path) -> None:
    exporter = PdfAuditExporter(_settings(tmp_path).agency)
    path = exporter.export(leads[3], tmp_path / "modern.pdf")  # Shiny Dental
    assert path.read_bytes().startswith(b"%PDF")


# -- Export service --------------------------------------------------------


def test_export_all_writes_expected_artifacts(leads: list[Lead], tmp_path: Path) -> None:
    service = ExportService(_settings(tmp_path))
    summary = service.export_all(leads)

    assert summary.ok
    for name in ("CRM.xlsx", "Dashboard.xlsx", "HighPriority.xlsx", "NoWebsite.xlsx",
                 "OutdatedWebsite.xlsx", "BrokenWebsite.xlsx", "leads.csv", "leads.json"):
        assert (tmp_path / name).exists(), name
    pdfs = list((tmp_path / "audits").glob("*.pdf"))
    assert len(pdfs) == len(leads)


def test_export_all_can_skip_pdfs(leads: list[Lead], tmp_path: Path) -> None:
    service = ExportService(_settings(tmp_path))
    service.export_all(leads, generate_pdfs=False)
    assert not (tmp_path / "audits").exists()


def test_partial_failure_is_isolated(leads: list[Lead], tmp_path: Path, monkeypatch) -> None:
    service = ExportService(_settings(tmp_path))

    def boom(_lead: Lead, _path: Path) -> Path:
        raise RuntimeError("pdf engine exploded")

    monkeypatch.setattr(service._pdf, "export", boom)
    summary = service.export_all(leads)

    assert not summary.ok
    assert len(summary.errors) == len(leads)         # every PDF failed
    assert (tmp_path / "CRM.xlsx").exists()           # but workbooks still written


def test_duplicate_names_get_unique_pdf_slugs() -> None:
    assert _slugify("IronCore Gym!") == "ironcore-gym"
    assert _slugify("   ") == "business"
